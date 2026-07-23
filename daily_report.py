#!/usr/bin/env python3
"""Daily Excel report of all companies notified about, emailed as an attachment.

Builds a bifurcated workbook (BSE 500 vs non-BSE 500) from seen.json — the
"New since last email" company sheets plus the all-time lists, sorted
earliest-first on first notification — and emails it via Gmail SMTP.
Days with nothing new send no email; those dates are listed in the next
email that does go out. State lives in reported.json (committed back).

Required env vars:
  GMAIL_ADDRESS      sender Gmail account
  GMAIL_APP_PASSWORD Gmail app password (not the account password)
  REPORT_EMAIL_TO    comma-separated recipients (defaults to GMAIL_ADDRESS)
"""
import json
import os
import re
import smtplib
import sys
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta, timezone
from email.message import EmailMessage
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
IST = timezone(timedelta(hours=5, minutes=30))
# High-water mark: NEWSIDs that have already gone out in an email. Committed back
# by the workflow so each email's "new" sheets = announcements added since the last.
REPORTED_PATH = ROOT / "reported.json"

DETAIL_API = "https://api.bseindia.com/BseIndiaAPI/api/CorpAnnouncementDTNewDataBeta/w"
PDF_BASE = "https://www.bseindia.com"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.bseindia.com/",
    "Origin": "https://www.bseindia.com",
}
SLUG_RE = re.compile(r"/stock-share-price/([^/]+)/")

HEAD_FILL = PatternFill("solid", fgColor="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
NON_FILL = PatternFill("solid", fgColor="843C0C")


def slug_name(nsurl, fallback):
    if nsurl:
        m = SLUG_RE.search(nsurl)
        if m:
            return m.group(1).replace("-", " ").strip().title()
    return fallback or ""


def fetch_one(session, bse500, bse500_set, nid):
    for attempt in range(3):
        try:
            r = session.get(DETAIL_API, params={"newsid": nid}, timeout=20)
            r.raise_for_status()
            rows = (r.json() or {}).get("Table") or []
            if not rows:
                # BSE sometimes returns an empty Table transiently; only conclude
                # the announcement is really gone after exhausting retries.
                if attempt == 2:
                    return {"NewsID": nid, "_status": "not_found"}
                time.sleep(1.5 * (attempt + 1))
                continue
            row = rows[0]
            scrip = str(row.get("Scrip_cd") or "").strip()
            pdf = row.get("PDF_Link") or ""
            return {
                "NewsID": nid,
                "scrip": scrip,
                "bse500": scrip in bse500_set,
                "company": bse500.get(scrip)
                or slug_name(row.get("NSUrl") or "", row.get("CompanyName") or ""),
                "category": (row.get("CATEGORYNAME") or "").strip(),
                "subject": (row.get("NewsSub") or "").strip(),
                "date": (row.get("News_dt") or "").replace("T", " ")[:19],
                "pdf": (PDF_BASE + pdf) if pdf else "",
                "_status": "ok",
            }
        except Exception as e:
            if attempt == 2:
                return {"NewsID": nid, "_status": f"error: {e}"}
            time.sleep(1.5 * (attempt + 1))


def aggregate(records):
    """Group per company; sorted earliest-first on first-notified time."""
    by = {}
    for r in records:
        key = r["scrip"] or f"name:{r['company']}"
        g = by.setdefault(
            key,
            {"company": r["company"], "scrip": r["scrip"], "count": 0, "first": None},
        )
        g["count"] += 1
        d = r["date"]
        if d:
            g["first"] = d if g["first"] is None or d < g["first"] else g["first"]
    return sorted(by.values(), key=lambda x: x["first"] or "9999")


def company_sheet(wb, title, companies, fill):
    s = wb.create_sheet(title)
    s.append(["Company", "Scrip Code", "Notifications", "First notified"])
    for c in s[1]:
        c.fill = fill
        c.font = HEAD_FONT
    for g in companies:
        s.append([g["company"], g["scrip"], g["count"], None])
        cell = s.cell(row=s.max_row, column=4)
        if g["first"]:
            # Real Excel date (numeric serial under the hood), not text.
            cell.value = datetime.strptime(g["first"][:10], "%Y-%m-%d").date()
            cell.number_format = "dd-mmm-yyyy"
    for i, w in enumerate([46, 12, 14, 16], 1):
        s.column_dimensions[get_column_letter(i)].width = w
    s.freeze_panes = "A2"
    for row in s.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def detail_sheet(wb, title, records):
    s = wb.create_sheet(title)
    s.append(["Group", "Company", "Scrip", "Category", "Date (IST)", "Subject", "PDF", "NewsID"])
    for c in s[1]:
        c.fill = HEAD_FILL
        c.font = HEAD_FONT
    for r in sorted(records, key=lambda x: x["date"], reverse=True):
        s.append([
            "BSE 500" if r["bse500"] else "Non-BSE 500",
            r["company"], r["scrip"], r["category"], r["date"], r["subject"], r["pdf"], r["NewsID"],
        ])
    for i, w in enumerate([13, 42, 10, 24, 20, 60, 58, 38], 1):
        s.column_dimensions[get_column_letter(i)].width = w
    s.freeze_panes = "A2"
    for row in s.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def load_report_state():
    """Return (reported_ids or None, quiet_days).

    reported_ids: NEWSIDs already covered by a sent email (None = first run).
    quiet_days:   display dates of suppressed no-news days since the last email.
    Handles the legacy plain-list file format transparently.
    """
    if REPORTED_PATH.exists():
        try:
            data = json.loads(REPORTED_PATH.read_text())
            if isinstance(data, dict):
                return set(data.get("ids") or []), list(data.get("quiet_days") or [])
            if isinstance(data, list):  # legacy format
                return set(data), []
        except Exception:
            return set(), []
    return None, []


def save_report_state(newsids, quiet_days):
    payload = {"ids": sorted(set(newsids)), "quiet_days": list(quiet_days)}
    REPORTED_PATH.write_text(json.dumps(payload, indent=2) + "\n")


def build_workbook(out_path):
    newsids = json.loads((ROOT / "seen.json").read_text())
    bse500 = json.loads((ROOT / "bse500.json").read_text())["constituents"]
    bse500_set = set(bse500)
    today_str = datetime.now(IST).strftime("%Y-%m-%d")
    reported, quiet_days = load_report_state()
    first_run = reported is None
    print(f"Resolving {len(newsids)} NEWSIDs (first_run={first_run})...", file=sys.stderr)

    session = requests.Session()
    session.headers.update(HEADERS)
    rows = []
    with ThreadPoolExecutor(max_workers=8) as ex:
        futs = {ex.submit(fetch_one, session, bse500, bse500_set, n): n for n in newsids}
        for f in as_completed(futs):
            rows.append(f.result())

    ok = [r for r in rows if r.get("_status") == "ok"]
    # "New since last email": announcements not yet emailed. On the very first
    # run (no high-water mark yet) fall back to today's date so the first email
    # isn't a giant backfill; tracking is exact from the second email onward.
    if first_run:
        new_rows = [r for r in ok if r["date"][:10] == today_str]
    else:
        new_rows = [r for r in ok if r["NewsID"] not in reported]

    bse_rows = [r for r in ok if r["bse500"]]
    non_rows = [r for r in ok if not r["bse500"]]
    bse_c, non_c = aggregate(bse_rows), aggregate(non_rows)
    new_bse_c = aggregate([r for r in new_rows if r["bse500"]])
    new_non_c = aggregate([r for r in new_rows if not r["bse500"]])

    wb = Workbook()
    company_sheet(wb, f"New BSE 500 ({len(new_bse_c)})", new_bse_c, HEAD_FILL)
    company_sheet(wb, f"New non-BSE 500 ({len(new_non_c)})", new_non_c, NON_FILL)
    company_sheet(wb, f"BSE 500 companies ({len(bse_c)})", bse_c, HEAD_FILL)
    company_sheet(wb, f"Non-BSE 500 companies ({len(non_c)})", non_c, NON_FILL)
    detail_sheet(wb, "All announcements", ok)
    wb.remove(wb["Sheet"])  # drop the default empty sheet

    wb.save(out_path)
    return {
        "new": len(new_rows),
        "new_companies": len(new_bse_c) + len(new_non_c),
        "total": len(ok),
        "companies": len(bse_c) + len(non_c),
        "bse500_companies": len(bse_c),
        "non_companies": len(non_c),
        "date": today_str,
        "first_run": first_run,
        "quiet_days": quiet_days,
        "prev_ids": sorted(reported) if reported is not None else [],
        # New high-water mark once the email actually sends. Rows whose detail
        # lookup errored are deliberately excluded so they retry next email
        # instead of being silently marked as already-reported.
        "all_ids": [r["NewsID"] for r in rows if r.get("_status") in ("ok", "not_found")],
    }


def send_email(xlsx_path, stats):
    sender = os.environ.get("GMAIL_ADDRESS")
    password = os.environ.get("GMAIL_APP_PASSWORD")
    to = os.environ.get("REPORT_EMAIL_TO") or sender
    if not sender or not password:
        print("GMAIL_ADDRESS / GMAIL_APP_PASSWORD not set", file=sys.stderr)
        sys.exit(1)

    msg = EmailMessage()
    msg["Subject"] = f"BSE Annual Report Watcher — daily report {stats['date']}"
    msg["From"] = sender
    msg["To"] = to
    quiet_line = ""
    if stats["quiet_days"]:
        quiet_line = (
            "No new annual report announcements on: "
            + ", ".join(stats["quiet_days"])
            + " (no email was sent on those days).\n\n"
        )
    msg.set_content(
        f"Daily report for {stats['date']} (IST).\n\n"
        f"{quiet_line}"
        f"New since last email: {stats['new']} announcements "
        f"across {stats['new_companies']} companies "
        f"(see the 'New ...' sheets).\n\n"
        f"Total to date: {stats['total']} announcements across "
        f"{stats['companies']} companies "
        f"({stats['bse500_companies']} BSE 500, {stats['non_companies']} non-BSE 500).\n\n"
        f"Full details in the attached Excel.\n\n"
        f"— Annual Report Watcher"
    )
    msg.add_attachment(
        Path(xlsx_path).read_bytes(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"bse-report-{stats['date']}.xlsx",
    )
    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=60) as smtp:
        smtp.login(sender, password)
        smtp.send_message(msg)
    print(f"Emailed {to}: new={stats['new']}, total={stats['total']}")


def main():
    out = ROOT / "daily_report.xlsx"
    stats = build_workbook(out)

    # Quiet day: nothing new since the last email. Send no email at all; just
    # bank today's date so the next real email can say which days were quiet.
    if stats["new"] == 0 and not stats["first_run"]:
        today_disp = datetime.strptime(stats["date"], "%Y-%m-%d").strftime("%d %b %Y")
        quiet = stats["quiet_days"] + [today_disp]
        save_report_state(stats["prev_ids"], quiet)
        print(f"No new announcements; email suppressed. {len(quiet)} quiet day(s) pending.")
        return

    send_email(out, stats)
    # Only advance the high-water mark (and clear the quiet-day list) after the
    # email has actually gone out, so a failed send retries the same set next run.
    save_report_state(stats["all_ids"], [])


if __name__ == "__main__":
    main()
